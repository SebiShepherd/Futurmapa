#!/usr/bin/env python3
"""Werkzeug zur Generierung der `scripts/data.js` aus Excel- oder CSV-Quellen.

Beispielaufrufe:
    python tools/xlsx_to_datajs.py --xlsx data/data-source.xlsx --output scripts/data.js
    python tools/xlsx_to_datajs.py --csv-dir data/csv-export --check-only

Die Eingabe kann eine XLSX-Arbeitsmappe oder ein Verzeichnis mit CSV-
Exporten der Tabellenblätter sein. Die Ausgabe wird mit zwei Leerzeichen
Einrückung erzeugt und entspricht der Struktur der bisherigen DATA_CONFIG.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl

EXPECTED_SHEETS: Sequence[str] = (
    "categories",
    "continents",
    "countries",
    "points",
    "org_metrics",
    "org_progress",
    "org_compare",
)


def log(level: str, message: str) -> None:
    """Gibt eine strukturierte Logzeile auf stdout oder stderr aus."""

    level_normalized = level.upper()
    stream = sys.stderr if level_normalized in {"ERROR", "WARNING"} else sys.stdout
    print(f"{level_normalized}: {message}", file=stream)


@dataclass
class Category:
    key: str
    label: str
    icon_id: str
    color_hex: str
    description: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        data: Dict[str, Any] = {
            "label": self.label,
            "iconId": self.icon_id,
            "color": self.color_hex,
        }
        if self.description:
            data["description"] = self.description
        return data


@dataclass
class Continent:
    name: str
    countries: List[str]
    description: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        data: Dict[str, Any] = {
            "countries": self.countries,
        }
        if self.description:
            data["description"] = self.description
        return data


@dataclass
class Point:
    point_id: str
    country_iso: str
    title: str
    category_key: str
    longitude: float
    latitude: float
    description: Optional[str] = None
    coming_soon: bool = False
    org_blocks: Dict[str, "OrgBlock"] = field(default_factory=dict)
    compare_block: Optional["CompareBlock"] = None

    def coordinates(self) -> List[float]:
        return [self.longitude, self.latitude]


@dataclass
class Country:
    iso_code: str
    name: str
    continent: str
    active: bool
    overview: Optional[str]
    points: List[Point] = field(default_factory=list)


@dataclass
class OrgBlock:
    organization: str
    summary: Optional[str] = None
    metrics: List[Dict[str, Any]] = field(default_factory=list)
    progress: List[Dict[str, Any]] = field(default_factory=list)

    def to_optional_dict(self) -> Optional[Dict[str, Any]]:
        data: Dict[str, Any] = {}
        if self.summary:
            data["summary"] = self.summary
        if self.metrics:
            data["metrics"] = self.metrics
        if self.progress:
            data["progress"] = self.progress
        return data or None


@dataclass
class CompareBlock:
    left_label: str
    right_label: str
    summary: Optional[str] = None
    metrics: List[Dict[str, Any]] = field(default_factory=list)

    def to_optional_dict(self) -> Optional[Dict[str, Any]]:
        data: Dict[str, Any] = {
            "leftLabel": self.left_label,
            "rightLabel": self.right_label,
        }
        if self.summary:
            data["summary"] = self.summary
        if self.metrics:
            data["metrics"] = self.metrics
        return data


def parse_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if text in {"true", "yes", "1"}:
        return True
    if text in {"false", "no", "0", ""}:
        return False
    raise ValueError(f"Kann Bool-Wert nicht interpretieren: {value!r}")


def load_tables(source: Path) -> Dict[str, List[Dict[str, Any]]]:
    if source.is_file():
        if source.suffix.lower() != ".xlsx":
            raise ValueError(f"Unterstützte Eingabe ist .xlsx oder Verzeichnis mit CSV-Dateien, nicht {source.suffix}")
        return load_from_xlsx(source)
    if source.is_dir():
        return load_from_csv_dir(source)
    raise ValueError(f"Pfad nicht gefunden: {source}")


def load_from_xlsx(path: Path) -> Dict[str, List[Dict[str, Any]]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    tables: Dict[str, List[Dict[str, Any]]] = {}
    for sheet_name in EXPECTED_SHEETS:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Tabellenblatt '{sheet_name}' fehlt in {path.name}")
        ws = workbook[sheet_name]
        tables[sheet_name] = list(iter_rows_with_header(ws.iter_rows(values_only=True)))
    return tables


def load_from_csv_dir(directory: Path) -> Dict[str, List[Dict[str, Any]]]:
    tables: Dict[str, List[Dict[str, Any]]] = {}
    for sheet_name in EXPECTED_SHEETS:
        csv_path = directory / f"{sheet_name}.csv"
        if not csv_path.exists():
            raise ValueError(f"CSV-Datei '{csv_path}' fehlt")
        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = [clean_row(dict(row)) for row in reader if any((row or {}).values())]
        tables[sheet_name] = rows
    return tables


def iter_rows_with_header(rows: Iterable[Sequence[Any]]) -> List[Dict[str, Any]]:
    iterator = iter(rows)
    try:
        header_row = next(iterator)
    except StopIteration:
        return []
    headers = [normalize_header(cell) for cell in header_row]
    result: List[Dict[str, Any]] = []
    for row in iterator:
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
            continue
        values = {
            headers[idx]: row[idx] if idx < len(row) else None
            for idx in range(len(headers))
        }
        result.append(clean_row(values))
    return result


def normalize_header(cell: Any) -> str:
    if cell is None:
        raise ValueError("Leere Spaltenüberschrift gefunden")
    return str(cell).strip()


def clean_row(row: Dict[str, Any]) -> Dict[str, Any]:
    cleaned: Dict[str, Any] = {}
    for key, value in row.items():
        if isinstance(value, str):
            cleaned[key] = value.strip()
        else:
            cleaned[key] = value
    return cleaned


def build_data(tables: Dict[str, List[Dict[str, Any]]]) -> Tuple[Optional[Dict[str, Any]], List[str]]:
    errors: List[str] = []
    categories = parse_categories(tables["categories"], errors)
    continents = parse_continents(tables["continents"], errors)
    countries = parse_countries(tables["countries"], continents, errors)
    points = parse_points(tables["points"], categories, countries, errors)
    validate_continent_country_links(continents, countries, errors)

    organization_order: List[str] = []
    org_blocks = parse_org_metrics(tables["org_metrics"], points, organization_order, errors)
    parse_org_progress(tables["org_progress"], points, org_blocks, organization_order, errors)
    compare_blocks = parse_org_compare(tables["org_compare"], points, errors)

    for point_id, point in points.items():
        point.compare_block = compare_blocks.get(point_id)
        if point_id in org_blocks:
            point.org_blocks = org_blocks[point_id]

    ensure_org_blocks_complete(points, organization_order, errors)

    if errors:
        return None, errors

    org_options = build_org_options(organization_order, compare_blocks)

    categories_dict = OrderedDict((cat.key, cat.to_dict()) for cat in categories)
    continents_dict = OrderedDict((cont.name, cont.to_dict()) for cont in continents)

    countries_dict: "OrderedDict[str, Dict[str, Any]]" = OrderedDict()
    for iso_code, country in countries.items():
        countries_dict[iso_code] = country_to_dict(country, organization_order)

    data_config = {
        "categories": categories_dict,
        "continents": continents_dict,
        "countries": countries_dict,
    }

    return {
        "org_options": org_options,
        "data_config": data_config,
    }, []


def parse_categories(rows: List[Dict[str, Any]], errors: List[str]) -> List[Category]:
    result: List[Category] = []
    seen: set[str] = set()
    for row in rows:
        key = require_field(row, "category_key", errors)
        if not key:
            continue
        if key in seen:
            errors.append(f"Kategorie-Schlüssel '{key}' ist doppelt vorhanden")
            continue
        seen.add(key)
        label = require_field(row, "label", errors, context=f"Kategorie '{key}'") or ""
        icon_id = require_field(row, "icon_id", errors, context=f"Kategorie '{key}'") or ""
        color_hex = require_field(row, "color_hex", errors, context=f"Kategorie '{key}'") or ""
        description = row.get("description") or None
        result.append(Category(key=key, label=label, icon_id=icon_id, color_hex=color_hex, description=description))
    return result


def parse_continents(rows: List[Dict[str, Any]], errors: List[str]) -> List[Continent]:
    result: List[Continent] = []
    seen: set[str] = set()
    for row in rows:
        name = require_field(row, "continent_name", errors)
        if not name:
            continue
        if name in seen:
            errors.append(f"Kontinent '{name}' ist mehrfach vorhanden")
            continue
        seen.add(name)
        iso_list_raw = require_field(row, "country_iso_list", errors, context=f"Kontinent '{name}'") or ""
        iso_list = [part.strip() for part in iso_list_raw.split(",") if part and part.strip()]
        description = row.get("description") or None
        result.append(Continent(name=name, countries=iso_list, description=description))
    return result


def parse_countries(
    rows: List[Dict[str, Any]],
    continents: List[Continent],
    errors: List[str],
) -> "OrderedDict[str, Country]":
    continent_names = {continent.name for continent in continents}
    result: "OrderedDict[str, Country]" = OrderedDict()
    for row in rows:
        iso = require_field(row, "country_iso3", errors)
        if not iso:
            continue
        if iso in result:
            errors.append(f"Land mit ISO-Code '{iso}' ist doppelt")
            continue
        name = require_field(row, "name", errors, context=f"Land '{iso}'") or ""
        continent = require_field(row, "continent_name", errors, context=f"Land '{iso}'") or ""
        if continent and continent not in continent_names:
            errors.append(f"Land '{iso}' referenziert unbekannten Kontinent '{continent}'")
        active = parse_bool_safe(row.get("active_flag"), errors, context=f"Land '{iso}'")
        overview = row.get("overview") or ""
        result[iso] = Country(iso_code=iso, name=name, continent=continent, active=active, overview=overview)
    return result


def parse_points(
    rows: List[Dict[str, Any]],
    categories: List[Category],
    countries: "OrderedDict[str, Country]",
    errors: List[str],
) -> Dict[str, Point]:
    category_keys = {category.key for category in categories}
    result: Dict[str, Point] = {}
    for row in rows:
        point_id = require_field(row, "point_id", errors)
        if not point_id:
            continue
        if point_id in result:
            errors.append(f"Marker-ID '{point_id}' ist doppelt")
            continue
        iso = require_field(row, "country_iso3", errors, context=f"Marker '{point_id}'") or ""
        if iso and iso not in countries:
            errors.append(f"Marker '{point_id}' verweist auf unbekanntes Land '{iso}'")
            continue
        title = require_field(row, "title", errors, context=f"Marker '{point_id}'") or ""
        category_key = require_field(row, "category_key", errors, context=f"Marker '{point_id}'") or ""
        if category_key and category_key not in category_keys:
            errors.append(f"Marker '{point_id}' verweist auf unbekannte Kategorie '{category_key}'")
        longitude = parse_float(row.get("longitude"), errors, f"Marker '{point_id}' (longitude)")
        latitude = parse_float(row.get("latitude"), errors, f"Marker '{point_id}' (latitude)")
        description = row.get("description") or None
        coming_soon = parse_bool_safe(row.get("coming_soon_flag"), errors, context=f"Marker '{point_id}'", default=False)
        point = Point(
            point_id=point_id,
            country_iso=iso,
            title=title,
            category_key=category_key,
            longitude=longitude,
            latitude=latitude,
            description=description,
            coming_soon=coming_soon,
        )
        result[point_id] = point
        if iso in countries:
            countries[iso].points.append(point)
    return result


def validate_continent_country_links(
    continents: List[Continent],
    countries: "OrderedDict[str, Country]",
    errors: List[str],
) -> None:
    country_isos = set(countries.keys())
    assigned_isos: set[str] = set()
    for continent in continents:
        for iso in continent.countries:
            if iso not in country_isos:
                errors.append(
                    f"Kontinent '{continent.name}' enthält unbekannten ISO-Code '{iso}'"
                )
            assigned_isos.add(iso)
    for iso in country_isos - assigned_isos:
        errors.append(f"Land '{iso}' ist keinem Kontinent in der ISO-Liste zugeordnet")


def parse_org_metrics(
    rows: List[Dict[str, Any]],
    points: Dict[str, Point],
    organization_order: List[str],
    errors: List[str],
) -> Dict[str, Dict[str, OrgBlock]]:
    blocks: Dict[str, Dict[str, OrgBlock]] = defaultdict(dict)
    for row in rows:
        point_id = require_field(row, "point_id", errors)
        if not point_id:
            continue
        if point_id not in points:
            errors.append(f"org_metrics: unbekannter Marker '{point_id}'")
            continue
        organization = require_field(row, "organization", errors, context=f"org_metrics für '{point_id}'")
        if not organization:
            continue
        if organization not in organization_order:
            organization_order.append(organization)
        block = blocks[point_id].get(organization)
        if block is None:
            block = OrgBlock(organization=organization)
            blocks[point_id][organization] = block
        summary = row.get("summary") or None
        if summary:
            if block.summary and block.summary != summary:
                errors.append(
                    f"org_metrics: widersprüchliche Zusammenfassung für Marker '{point_id}', Organisation '{organization}'"
                )
            else:
                block.summary = summary
        metric_label = (row.get("metric_label") or "").strip()
        metric_value = row.get("metric_value")
        metric_trend = row.get("metric_trend")
        if metric_label or metric_value or metric_trend:
            if not metric_label or metric_value in (None, ""):
                errors.append(
                    f"org_metrics: unvollständige Kennzahl für Marker '{point_id}', Organisation '{organization}'"
                )
            else:
                metric_entry: Dict[str, Any] = {
                    "label": metric_label,
                    "value": str(metric_value),
                }
                if metric_trend not in (None, ""):
                    metric_entry["trend"] = str(metric_trend)
                block.metrics.append(metric_entry)
    return blocks


def parse_org_progress(
    rows: List[Dict[str, Any]],
    points: Dict[str, Point],
    org_blocks: Dict[str, Dict[str, OrgBlock]],
    organization_order: List[str],
    errors: List[str],
) -> None:
    for row in rows:
        point_id = require_field(row, "point_id", errors)
        if not point_id:
            continue
        if point_id not in points:
            errors.append(f"org_progress: unbekannter Marker '{point_id}'")
            continue
        organization = require_field(row, "organization", errors, context=f"org_progress für '{point_id}'")
        if not organization:
            continue
        if organization not in organization_order:
            organization_order.append(organization)
        block = org_blocks.setdefault(point_id, {}).setdefault(organization, OrgBlock(organization=organization))
        summary = row.get("summary") or None
        if summary:
            if block.summary and block.summary != summary:
                errors.append(
                    f"org_progress: Zusammenfassung für Marker '{point_id}', Organisation '{organization}' widerspricht org_metrics"
                )
            else:
                block.summary = summary
        progress_label = (row.get("progress_label") or "").strip()
        progress_value = row.get("progress_value")
        if progress_label or progress_value not in (None, ""):
            if not progress_label or progress_value in (None, ""):
                errors.append(
                    f"org_progress: unvollständiger Fortschrittswert für Marker '{point_id}', Organisation '{organization}'"
                )
            else:
                block.progress.append({"label": progress_label, "value": progress_value})


def parse_org_compare(
    rows: List[Dict[str, Any]],
    points: Dict[str, Point],
    errors: List[str],
) -> Dict[str, CompareBlock]:
    compares: Dict[str, CompareBlock] = {}
    for row in rows:
        point_id = require_field(row, "point_id", errors)
        if not point_id:
            continue
        if point_id not in points:
            errors.append(f"org_compare: unbekannter Marker '{point_id}'")
            continue
        left_label = require_field(row, "left_label", errors, context=f"org_compare für '{point_id}'")
        right_label = require_field(row, "right_label", errors, context=f"org_compare für '{point_id}'")
        if not left_label or not right_label:
            continue
        block = compares.get(point_id)
        if block is None:
            block = CompareBlock(left_label=left_label, right_label=right_label)
            compares[point_id] = block
        else:
            if block.left_label != left_label or block.right_label != right_label:
                errors.append(
                    f"org_compare: inkonsistente Vergleichslabels für Marker '{point_id}'"
                )
        summary = row.get("summary") or None
        if summary:
            if block.summary and block.summary != summary:
                errors.append(f"org_compare: widersprüchliche Zusammenfassung für Marker '{point_id}'")
            else:
                block.summary = summary
        metric_label = (row.get("metric_label") or "").strip()
        left_value = row.get("left_value")
        right_value = row.get("right_value")
        if metric_label or left_value not in (None, "") or right_value not in (None, ""):
            if not metric_label or left_value in (None, "") or right_value in (None, ""):
                errors.append(f"org_compare: unvollständige Vergleichskennzahl für Marker '{point_id}'")
            else:
                block.metrics.append({
                    "label": metric_label,
                    "left": str(left_value),
                    "right": str(right_value),
                })
    return compares


def ensure_org_blocks_complete(
    points: Dict[str, Point],
    organization_order: List[str],
    errors: List[str],
) -> None:
    expected = set(organization_order)
    for point in points.values():
        available = set(point.org_blocks.keys())
        missing = expected - available
        if missing and not point.coming_soon:
            errors.append(
                f"Marker '{point.point_id}' hat fehlende Organisationsdaten für: {', '.join(sorted(missing))}"
            )
        if point.coming_soon:
            has_real_data = any(
                block.to_optional_dict() is not None for block in point.org_blocks.values()
            ) or (point.compare_block and point.compare_block.metrics)
            if has_real_data:
                errors.append(
                    f"Marker '{point.point_id}' ist als Coming Soon markiert, enthält aber Kennzahlen"
                )


def parse_bool_safe(value: Any, errors: List[str], context: str, default: bool = False) -> bool:
    try:
        return parse_bool(value, default=default)
    except ValueError as exc:
        errors.append(f"{context}: {exc}")
        return default


def parse_float(value: Any, errors: List[str], context: str) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    if value is None or (isinstance(value, str) and value.strip() == ""):
        errors.append(f"{context}: Wert fehlt")
        return 0.0
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        errors.append(f"{context}: '{value}' ist keine gültige Zahl")
        return 0.0


def require_field(row: Dict[str, Any], key: str, errors: List[str], context: Optional[str] = None) -> Optional[str]:
    value = row.get(key)
    if value in (None, ""):
        prefix = f"{context}: " if context else ""
        errors.append(f"{prefix}Pflichtfeld '{key}' fehlt")
        return None
    return str(value)


def point_to_dict(point: Point, organization_order: Sequence[str]) -> Dict[str, Any]:
    data: Dict[str, Any] = {
        "id": point.point_id,
        "title": point.title,
        "category": point.category_key,
        "coordinates": point.coordinates(),
    }
    if point.description:
        data["description"] = point.description
    if point.coming_soon:
        data["comingSoon"] = True
        return data
    if point.org_blocks:
        data_block: "OrderedDict[str, Any]" = OrderedDict()
        for organization in organization_order:
            block = point.org_blocks.get(organization)
            data_block[organization] = block.to_optional_dict() if block else None
        compare_dict = point.compare_block.to_optional_dict() if point.compare_block else None
        data_block["compare"] = compare_dict if compare_dict and compare_dict.get("metrics") else None
        data["data"] = data_block
    elif point.compare_block:
        compare_dict = point.compare_block.to_optional_dict()
        data["data"] = OrderedDict([("compare", compare_dict if compare_dict and compare_dict.get("metrics") else None)])
    return data


def country_to_dict(country: Country, organization_order: Sequence[str]) -> Dict[str, Any]:
    return {
        "name": country.name,
        "continent": country.continent,
        "active": country.active,
        "overview": country.overview or "",
        "points": [point_to_dict(point, organization_order) for point in country.points],
    }


def build_org_options(organization_order: List[str], compare_blocks: Dict[str, CompareBlock]) -> List[str]:
    options = list(organization_order)
    for block in compare_blocks.values():
        label = f"{block.left_label} vs {block.right_label}"
        if label not in options:
            options.append(label)
    return options


def render_js(data: Dict[str, Any]) -> str:
    org_options_json = json.dumps(data["org_options"], ensure_ascii=False, indent=2)
    data_config_json = json.dumps(data["data_config"], ensure_ascii=False, indent=2)
    lines = [
        "// Auto-generated by tools/xlsx_to_datajs.py – DO NOT EDIT",
        f"const ORG_OPTIONS = {org_options_json};",
        "",
        f"const DATA_CONFIG = {data_config_json};",
        "",
        "const COUNTRY_BY_ISO = new Map(Object.entries(DATA_CONFIG.countries));",
        "const CONTINENT_LIST = Object.keys(DATA_CONFIG.continents);",
        "",
    ]
    return "\n".join(lines)


def write_output(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write(content)


def parse_arguments(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Konvertiert Datenquellen in scripts/data.js")
    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument("--xlsx", metavar="PFAD", help="Pfad zur XLSX-Datei mit allen Tabellenblättern")
    input_group.add_argument(
        "--csv-dir",
        metavar="VERZEICHNIS",
        help="Verzeichnis mit CSV-Dateien (eine Datei pro Tabellenblatt)",
    )
    parser.add_argument(
        "--output",
        "-o",
        metavar="DATEI",
        help="Zieldatei (JavaScript)",
    )
    parser.add_argument(
        "--check-only",
        action="store_true",
        help="Nur Validierung durchführen, keine Datei schreiben",
    )
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_arguments(argv)

    if not args.check_only and not args.output:
        log("ERROR", "--output ist erforderlich, wenn nicht --check-only genutzt wird")
        return 2

    try:
        if args.xlsx:
            source_path = Path(args.xlsx)
            tables = load_from_xlsx(source_path)
            source_description = f"XLSX-Datei {source_path}"
        else:
            source_path = Path(args.csv_dir)
            tables = load_from_csv_dir(source_path)
            source_description = f"CSV-Verzeichnis {source_path}"
    except Exception as exc:  # noqa: BLE001
        log("ERROR", f"Quelldaten konnten nicht geladen werden: {exc}")
        return 2

    log("INFO", f"Quelldaten erfolgreich gelesen aus {source_description}")

    data, errors = build_data(tables)
    if errors:
        for message in errors:
            log("ERROR", message)
        log("ERROR", "Validierung fehlgeschlagen.")
        return 1

    assert data is not None

    if args.check_only:
        log("INFO", "Validierung erfolgreich abgeschlossen (Check-Only).")
        return 0

    output_path = Path(args.output)
    try:
        content = render_js(data)
        write_output(output_path, content)
    except Exception as exc:  # noqa: BLE001
        log("ERROR", f"Ausgabe konnte nicht geschrieben werden: {exc}")
        return 3

    log("INFO", f"Datei '{output_path}' aktualisiert.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
